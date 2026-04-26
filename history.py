import os
import sqlite3
from datetime import datetime, timezone
from dateutil import parser as dtparser
from openpyxl import load_workbook

try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None


DB_PATH = os.getenv("DB_PATH", "/var/data/app.db")
EXCEL_PATH = "history.xlsx"
DEVICE = os.getenv("IMPORT_DEVICE", "pi4")
LOCAL_TZ_NAME = os.getenv("IMPORT_TZ", "Asia/Manila")

NOTE_METRIC_CANONICAL = "real_power"
NOTE_METRIC_LEGACY = "power"

FIELD_ALIASES = {
    "time": ["time", "timestamp", "date time", "datetime"],
    "rms_voltage": ["voltage", "voltage (v)", "rms_voltage", "v"],
    "rms_current": ["current", "current (a)", "rms_current", "a"],
    "power": ["power", "power (w)", "p"],
    "power_factor": ["power factor", "power_factor", "pf"],
    "note": ["note", "notes", "remarks", "remark"],
}


def clean_header(value):
    return str(value or "").strip().lower()


def clean_text(value):
    if value is None:
        return ""

    text = str(value).strip()

    if text.lower() in ["nan", "none", "null"]:
        return ""

    return text


def parse_number(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()

    if not text or text in ["—", "-"]:
        return None

    text = (
        text.replace(",", "")
        .replace(" V", "")
        .replace(" A", "")
        .replace(" W", "")
        .replace("v", "")
        .replace("a", "")
        .replace("w", "")
        .strip()
    )

    if not text:
        return None

    try:
        return float(text)
    except Exception:
        return None


def get_local_tz():
    if ZoneInfo is None:
        return timezone.utc

    try:
        return ZoneInfo(LOCAL_TZ_NAME)
    except Exception:
        return timezone.utc


def parse_time_value(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    text = str(value).strip()

    if not text:
        return None

    try:
        return dtparser.parse(text)
    except Exception:
        return None


def display_key_from_dt(dt):
    if dt is None:
        return ""

    hour = dt.hour % 12

    if hour == 0:
        hour = 12

    ampm = "AM" if dt.hour < 12 else "PM"

    return f"{dt.month}/{dt.day}/{dt.year}, {hour}:{dt.minute:02d}:{dt.second:02d} {ampm}"


def display_key_from_any(value):
    dt = parse_time_value(value)

    if dt is None:
        return clean_text(value)

    return display_key_from_dt(dt)


def parse_db_time(db_time):
    try:
        return dtparser.parse(str(db_time).replace("Z", "+00:00"))
    except Exception:
        return None


def configure_sqlite(conn):
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = FULL")
    conn.execute("PRAGMA foreign_keys = ON")


def build_db_time_lookup(conn, device):
    local_tz = get_local_tz()

    rows = conn.execute(
        """
        SELECT DISTINCT time
        FROM realtime_points
        WHERE device=?
        """,
        (device,),
    ).fetchall()

    lookup = {}

    for row in rows:
        db_time = str(row["time"]).strip()

        if db_time:
            lookup[db_time] = db_time

        dt = parse_db_time(db_time)

        if dt is None:
            continue

        if dt.tzinfo is not None:
            local_dt = dt.astimezone(local_tz)
            lookup[display_key_from_dt(local_dt)] = db_time

            utc_dt = dt.astimezone(timezone.utc)
            lookup[display_key_from_dt(utc_dt)] = db_time
        else:
            lookup[display_key_from_dt(dt)] = db_time

            utc_dt = dt.replace(tzinfo=timezone.utc)
            local_dt = utc_dt.astimezone(local_tz)
            lookup[display_key_from_dt(local_dt)] = db_time

    return lookup


def read_excel_rows(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active

    header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [clean_header(header) for header in header_values]

    rows = []

    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {}

        for index, header in enumerate(headers):
            row[header] = values[index] if index < len(values) else None

        rows.append(row)

    return rows


def find_value(row, key):
    aliases = FIELD_ALIASES[key]

    for alias in aliases:
        if alias in row:
            return row[alias]

    return None


def get_author_id(conn):
    row = conn.execute("SELECT id FROM users ORDER BY id ASC LIMIT 1").fetchone()

    if row:
        return int(row["id"])

    return 0


def update_realtime(conn, device, field, db_time, new_value):
    if new_value is None:
        return False, None, None

    row = conn.execute(
        """
        SELECT id, value
        FROM realtime_points
        WHERE device=? AND field=? AND time=?
        ORDER BY id ASC
        LIMIT 1
        """,
        (device, field, db_time),
    ).fetchone()

    if row:
        old_value = float(row["value"])
        new_float = float(new_value)

        if abs(old_value - new_float) < 0.0000001:
            return False, old_value, new_float

        conn.execute(
            """
            UPDATE realtime_points
            SET value=?
            WHERE id=?
            """,
            (new_float, int(row["id"])),
        )

        return True, old_value, new_float

    conn.execute(
        """
        INSERT INTO realtime_points (device, field, time, value)
        VALUES (?, ?, ?, ?)
        """,
        (device, field, db_time, float(new_value)),
    )

    return True, None, float(new_value)


def upsert_note(conn, device, db_time, note_text, author_id, clear_blank_notes=False):
    existing = conn.execute(
        """
        SELECT *
        FROM notes
        WHERE device=?
          AND COALESCE(anchor_time, time)=?
          AND metric IN (?, ?)
        ORDER BY id ASC
        LIMIT 1
        """,
        (
            device,
            db_time,
            NOTE_METRIC_CANONICAL,
            NOTE_METRIC_LEGACY,
        ),
    ).fetchone()

    note_text = clean_text(note_text)

    if not note_text:
        if existing and clear_blank_notes:
            conn.execute("DELETE FROM notes WHERE id=?", (int(existing["id"]),))
            return "deleted"

        return "unchanged"

    now = datetime.now(timezone.utc).isoformat()

    power_row = conn.execute(
        """
        SELECT value
        FROM realtime_points
        WHERE device=? AND field='power' AND time=?
        ORDER BY id ASC
        LIMIT 1
        """,
        (device, db_time),
    ).fetchone()

    anchor_value = float(power_row["value"]) if power_row else None
    verified = 1 if power_row else 0

    if existing:
        if existing["text"] == note_text:
            return "unchanged"

        conn.execute(
            """
            UPDATE notes
            SET
                metric=?,
                time=?,
                text=?,
                updated_at=?,
                anchor_time=?,
                anchor_value=?,
                anchor_field=?,
                verified=?
            WHERE id=?
            """,
            (
                NOTE_METRIC_CANONICAL,
                db_time,
                note_text,
                now,
                db_time,
                anchor_value,
                "power",
                int(verified),
                int(existing["id"]),
            ),
        )

        return "updated"

    conn.execute(
        """
        INSERT INTO notes (
            device,
            metric,
            time,
            text,
            author_id,
            created_at,
            updated_at,
            anchor_time,
            anchor_value,
            anchor_field,
            verified
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            device,
            NOTE_METRIC_CANONICAL,
            db_time,
            note_text,
            int(author_id),
            now,
            now,
            db_time,
            anchor_value,
            "power",
            int(verified),
        ),
    )

    return "inserted"


def run_import(dry_run=False, clear_blank_notes=False):
    if not os.path.exists(EXCEL_PATH):
        raise SystemExit(f"Excel file not found: {EXCEL_PATH}")

    if not os.path.exists(DB_PATH):
        raise SystemExit(f"Database not found: {DB_PATH}")

    rows = read_excel_rows(EXCEL_PATH)

    conn = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=5.0)
    configure_sqlite(conn)

    matched_rows = 0
    skipped_rows = 0
    changed_values = 0
    inserted_notes = 0
    updated_notes = 0
    deleted_notes = 0
    unchanged_notes = 0
    examples = []

    try:
        lookup = build_db_time_lookup(conn, DEVICE)
        author_id = get_author_id(conn)

        for index, row in enumerate(rows, start=2):
            raw_time = find_value(row, "time")
            key = display_key_from_any(raw_time)

            db_time = lookup.get(key) or lookup.get(clean_text(raw_time))

            if not db_time:
                skipped_rows += 1

                if len(examples) < 10:
                    examples.append(
                        f"Skipped row {index}: no DB match for time {raw_time} / key {key}"
                    )

                continue

            matched_rows += 1

            values = {
                "rms_voltage": parse_number(find_value(row, "rms_voltage")),
                "rms_current": parse_number(find_value(row, "rms_current")),
                "power": parse_number(find_value(row, "power")),
                "power_factor": parse_number(find_value(row, "power_factor")),
            }

            for field, new_value in values.items():
                changed, old_value, saved_value = update_realtime(
                    conn,
                    DEVICE,
                    field,
                    db_time,
                    new_value,
                )

                if changed:
                    changed_values += 1

                    if len(examples) < 10:
                        examples.append(
                            f"{key} | {field}: {old_value} -> {saved_value}"
                        )

            note_value = find_value(row, "note")
            note_action = upsert_note(
                conn,
                DEVICE,
                db_time,
                note_value,
                author_id,
                clear_blank_notes=clear_blank_notes,
            )

            if note_action == "inserted":
                inserted_notes += 1
            elif note_action == "updated":
                updated_notes += 1
            elif note_action == "deleted":
                deleted_notes += 1
            else:
                unchanged_notes += 1

        if dry_run:
            conn.rollback()
        else:
            conn.commit()

        print("")
        print("Import check complete." if dry_run else "Import complete.")
        print(
            f"Mode: {'DRY RUN - no database changes saved' if dry_run else 'LIVE - database changes saved'}"
        )
        print(f"Database: {DB_PATH}")
        print(f"Input file: {EXCEL_PATH}")
        print(f"Device: {DEVICE}")
        print(f"Local timezone used for matching: {LOCAL_TZ_NAME}")
        print(f"Input rows: {len(rows)}")
        print(f"Matched rows: {matched_rows}")
        print(f"Skipped rows: {skipped_rows}")
        print(f"Realtime values changed: {changed_values}")
        print(f"Notes inserted: {inserted_notes}")
        print(f"Notes updated: {updated_notes}")
        print(f"Notes deleted: {deleted_notes}")
        print(f"Notes unchanged: {unchanged_notes}")

        if examples:
            print("")
            print("Examples:")

            for item in examples:
                print(" - " + item)

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    import sys

    dry_run = "--dry-run" in sys.argv
    clear_blank_notes = "--clear-blank-notes" in sys.argv

    run_import(
        dry_run=dry_run,
        clear_blank_notes=clear_blank_notes,
    )
